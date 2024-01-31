# Plot a pdf report file 

from email import charset
import io
import logging
from pathlib import Path
import tempfile
from fpdf import FPDF
from PIL import Image
import pandas as pd
import numpy as np
from alpinac_gui.matchms_funcs import formula_from_namestring, formula_to_latex, halocarbname2formula, plot_mass_spectra, load_from_mgf, metadata_processing, peak_processing, AlpinacData, halocarbname2formula
from pyvalem.formula import Formula
from molmass import Formula as chem_formula
import pubchempy as pcp
from matplotlib import pyplot as plt
from matplotlib import axes
from db_reader import AlpinacFragmentsReader, JdxReader, AbstractReader
from matchms.Spectrum import Spectrum
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
import chardet
import cirpy
from data_analysis_utils import formula_by_name, clean_empa_name_string

# FUNCTIONS

def get_elem(updated_header, row, i, pdf):
    elem = ""
    fill_opt = 0
    formula = ""
    case = {
        "0_score": lambda: handle_score_case(elem, fill_opt, row, i, pdf),
        "0_matching_cmp": lambda: handle_matching_cmp_case(elem, fill_opt, row, i, pdf),
        "alpinac_results": lambda: handle_alpinac_results_case(elem, fill_opt, row, i, pdf),
        "alpinac_results_no_N": lambda: handle_alpinac_results_case(elem, fill_opt, row, i, pdf),
        "alpinac_results_EI_CI": lambda: handle_alpinac_results_case(elem, fill_opt, row, i, pdf),
        "#peaks": lambda: handle_no_peaks_case(elem, fill_opt, row, i, pdf),
        "default": lambda: handle_default_case(elem, fill_opt, row, i, pdf)
    }
    # get the function for the current header or the default function
    func = case.get(str(updated_header), case["default"])
    # execute the function
    elem, fill_opt, formula = func()
    return elem, fill_opt, formula 
    
def handle_score_case(elem, fill_opt, row, i, pdf):
    #nonlocal elem, fill_opt
    formula = ""
    elem = row.split("\n")[i]
    if float(elem)>=0.9:
        pdf.set_fill_color(120, 240, 40)
        fill_opt = 1
    else:
        pdf.set_fill_color(255, 255, 255)
        fill_opt = 0
    return elem, fill_opt, formula

def handle_matching_cmp_case(elem, fill_opt, row, i, pdf, formulas_dict = {}):
    #nonlocal elem, fill_opt
    global formulas_ms
    elem = row.split("\n")[i]
    try:
        formulas_ms[i], _, _, _ , _, formulas_dict= formula_by_name(elem, formulas_dict, formulas_dict_manually_added, target_list)
    except:
        formula = ""
    print(formulas_ms)
    return elem, fill_opt, formula, formulas_dict
      
def handle_alpinac_results_case(elem, fill_opt, row, i, pdf):
    #nonlocal elem, fill_opt
    formula = ""
    try:
        # count the number of occurences of the character ";" in the string
        count = row.count(";")
        if count <= 1:
            elem = (row.replace(")",")\n").split("\n", maxsplit = 3)[i]).replace(";", "").strip()
            # print the count and the elem
            #print("count is : {}".format(count))
            #print("elem is: {}".format(elem))
            #print(row)
        else:
            elem = row.split(" ; ", maxsplit = 3)[i].replace(";", "")
    except:
        elem = ""
    # if part of string is a formula convert into lattex
    if len(elem) > 1:
        formula_str = formula_from_namestring(elem)
        #formula_ltx = formula_to_latex(formula_str.formula)
        #print(formula_str)
        if not formula_str:
                    elem = ""
                    print("formula could not be parsed")
        else:
            if str(chem_formula(str(formula_str)).formula) in formulas_ms and len(str(formula_str))>0:
                pdf.set_fill_color(120, 240, 40)
                fill_opt = 1
            else:
                pdf.set_fill_color(255, 255, 255)
                fill_opt = 0
            formula = str(chem_formula(str(formula_str)).formula)
            elem = elem.replace(str(formula_str), formula)
    return elem, fill_opt, formula

def handle_no_peaks_case(elem, fill_opt, row, i, pdf):
    #nonlocal elem, fill_opt
    formula = ""
    if i == 1:
        elem = row
    if int(row) <= 5:
        pdf.set_fill_color(240, 120, 40)
        fill_opt = 1
    else:
        pdf.set_fill_color(255, 255, 255)
        fill_opt = 0
    return elem, fill_opt, formula
       
def handle_default_case(elem, fill_opt, row, i, pdf):
    #nonlocal elem, fill_opt
    formula = ""
    if i == 1:
        elem = row
    return elem, fill_opt, formula

def add_matplotlib_plot_to_pdf_old(fig, pdf):
    # Convert plot to image
    buf = io.BytesIO()
    canvas = FigureCanvas(fig)
    canvas.print_figure(buf, format='png')
    buf.seek(0)
    pil_image = Image.open(buf)

    # Add image to PDF
    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
        pil_image.save(tmp_file.name)
        tmp_file.seek(0)
        image_path = tmp_file.name

        if pdf.get_y() + pil_image.height > pdf.h:
            pdf.add_page('P')

        if pil_image.width > 20000:
            pil_image = pil_image.resize((200, 100), Image.ANTIALIAS)

        pdf.image(image_path, x=0, y=pdf.get_y(), w=200, h=100)
        tmp_file.close()

def add_matplotlib_plot_to_pdf(fig, pdf):
    # Convert plot to image
    buf = io.BytesIO()
    canvas = FigureCanvas(fig)
    canvas.print_figure(buf, format='png')
    buf.seek(0)
    pil_image = Image.open(buf)

    # Get current y position and add image to PDF
    current_y = pdf.get_y()
    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
        pil_image.save(tmp_file.name)
        tmp_file.seek(0)
        image_path = tmp_file.name

        if current_y + pil_image.height > pdf.h:
            pdf.add_page('L')
            current_y = pdf.get_y()

        if pil_image.width > 20000:
            pil_image = pil_image.resize((200, 100), Image.ANTIALIAS)

        pdf.image(image_path, x=0, y=current_y, w=pdf.w)
        # set y position to the bottom of the image
        pdf.set_y(current_y + pil_image.height)
        tmp_file.close()

def plot_header_and_corresponding_line(row_ind):

    for index, col in enumerate(updated_header):
        print(index)
        pdf.cell(col_width_list[index], row_height, str(col).replace("0_", "cs_"), border=1, align = 'C')
        if col == "0_matching_cmp":
            pdf.cell(col_width_list[index]/2, row_height, 'formula', border=1, align="C")
    pdf.ln()

    # loop over the data and plot each cell
    #
    row = [str(table_data.iloc[row_ind][key]) for key in updated_header]

    for i in range(3):
        if i==2:
            border_val = 'BLR'
        else:
            border_val = 'LR'
        for col in range(len(row)):
            elem, fill_opt, formula = get_elem(updated_header[col], row[col], i, pdf)
            print(elem)
            pdf.cell(col_width_list[col], row_height, elem, border=border_val, align="C", fill=fill_opt)
            fill_opt = 0
            if str(updated_header[col]) == "0_matching_cmp":
                pdf.cell(col_width_list[col]/2, row_height, formula, border=border_val, align="C")
        pdf.ln(row_height)
        print(formulas_ms)

def find_block(ind, start_rows, len_table):
    # find block of rows where ind is in
    # start_rows: list of row indices where a new block starts
    # ind: index of row to be found in which block
    # returns: start and end index of block
    for n_start in start_rows:
        try:
            n_end = start_rows[np.where(start_rows > n_start)[0]][0]
        except:
            n_end = len_table
        if ind >= n_start and ind < n_end:
            return n_start, n_end
    return None

def check_if_subfragement(potential_subfragment, formula):
    # check if formula is a subfragment of another formula
    # e.g. C1H4 is a subfragment of C1H2
    # check if formulas are identical
    if chem_formula(potential_subfragment).formula == chem_formula(formula).formula:
        return False # is identical

    df_pot_sub = chem_formula(potential_subfragment).composition().dataframe()
    df_pot_sub['Count']
    df_formula = chem_formula(formula).composition().dataframe()
    df_formula['Count']
    # extend df_pot_sub and df_formula with missing elements with same relative mass but fraction and count = 0
    # get all elements in df_pot_sub and df_formula
    all_elements = np.unique(np.append(df_pot_sub.index.values,df_formula.index.values))
    # if not already in df_pot_sub or df_formula add element with fraction = 0 and count = 0
    for elem in all_elements:
        if elem not in df_pot_sub.index.values:
            df_pot_sub.loc[elem, 'Fraction'] = 0
            df_pot_sub.loc[elem, 'Count'] = 0
        if elem not in df_formula.index.values:
            df_formula.loc[elem, 'Fraction'] = 0
            df_formula.loc[elem, 'Count'] = 0

    
    # get row names
    #join df_pot_sub.index.values and df_formula.index.values
    all_rownames =np.unique(np.append(df_pot_sub.index.values,df_formula.index.values))
    cond_matched = []
    for elem in all_rownames:
        # get count of element in potential_subfragment
        try:
            count_pot_sub = df_pot_sub.loc[elem, 'Count']
        except:
            count_pot_sub = 0
        # get count of element in formula
        count_formula = df_formula.loc[elem, 'Count']
        if count_pot_sub <= count_formula:
            cond_matched.append(True)
        else:
            cond_matched.append(False)
    if all(cond_matched):
        return True
    else:
        return False

def find_alpinac_matching_pairs(n_ind, data, data_col_alp):
    """find all rows in data[data_col_alp] where the formula is identical to the formula in data[data_col_alp][n_ind] and 
        return the row indices of the matching rows
    """
    row_rel_matched = [] 
    row_subf = []  
    # get block of rows
    n_start, n_end = find_block(ind=n_ind, start_rows=row_ind_not_empty, len_table=len(data))
    
    try:
        formula_alp = data[data_col_alp][n_ind].split(':')[-1].split('(')[0].strip()
        candidates_cs =data['cs_formula'][n_start:n_end]
        #candidates_alp = [cand.split(':')[-1].split('(')[0].strip() for cand in candidates_alp_full_str]
        # if formula_cs is in new_table_data['cs_matching_cmp'][n_ind:n_next_row_start-1] set background color to green
        # test if formula_alp is a chemical formula
        formula_alp = chem_formula(formula_alp).formula
        valid_cs_formulas = [chem_formula(form).formula for form in candidates_cs if form != ""]
        ind_rel_matched = [ind for ind in range(len(valid_cs_formulas)) if valid_cs_formulas[ind] == formula_alp]
        ind_subf = [ind for ind in range(len(valid_cs_formulas)) if check_if_subfragement(potential_subfragment=formula_alp, formula=valid_cs_formulas[ind])]
        if ind_rel_matched:
            row_rel_matched = n_start+ind_rel_matched 
        if ind_subf:
            row_subf = n_start+ind_subf
    except:
        pass
    return row_rel_matched, row_subf    

def add_cmp_str(elem1, elem2):
    if elem1 == "" or elem2 == "":
        return_val = ""
    else:
        ncpm_in_brackets = ["("+subelem+")" for subelem in elem2.replace("[","").replace("]","").split(', ')]
        cmp_ids = elem1.replace("[", "").replace("]","").strip().split()
        cmp_ids = [cmp_id.strip() for cmp_id in cmp_ids]
        # convert to list if not already a list
        if not cmp_ids or cmp_ids == ['']:
            return_val = ""
        else:
            return_val = str(np.core.defchararray.add(np.array(cmp_ids), np.array(ncpm_in_brackets))).replace(" ", ", ").replace("'",'').replace('[','').replace(']','').strip()
    return return_val

def get_colour_table_old(data_summary):



    


    # generate colour_table with same dimensions as table_data, filled with white in hex notation
    color_hex_table = pd.DataFrame('#00FFFFFF', index=data_summary.index, columns=data_summary.columns)

    # GREY OUT UNUSABLE DATA PEAKS
    col_no = col_greyed_out
    ind_greyed_out = []
    for n_ind in row_ind_not_empty:
        if (data_summary['alp. cmp(peaks) EI'][n_ind].startswith('no') and data_summary['alp. cmp(peaks) EI'][n_ind].startswith('no')) or int(data_summary['#peaks'][n_ind]) < 5:
            n_start, n_end = find_block(ind=n_ind, start_rows=row_ind_not_empty, len_table=len(data_summary))
            # add all values between n_start and n_end ind_greyed_out
            ind_greyed_out.extend(range(n_start, n_end))
            data_summary['eval_status'][n_ind] = 'no usable data'
    color_hex_table.values[ind_greyed_out] = col_no

    # GREEN BLOCK IF MATCHING SCORE IS HIGH ENOUGH

    ind_validated = []
    ind_valid_high_prop = []
    for n_ind in range(len(data_summary['cs_score'])):
        if pd.to_numeric(data_summary['cs_score'][n_ind], errors='coerce') >= lim_ok_score:
            n_start, n_end = find_block(ind=n_ind, start_rows=row_ind_not_empty, len_table=len(data_summary))
            # if also a matching pair is found, set the color to light green
            dict_ind_rel_matched = {}
            dict_ind_subfragment = {}
            for alp_col_str in ['alpinac_EI', 'alpinac_EI_no_N', 'alpinac_CI', 'alpinac_CI_no_N']:
                dict_ind_rel_matched[alp_col_str], dict_ind_subfragment[alp_col_str] = find_alpinac_matching_pairs(n_ind=n_ind, data=data_summary, data_col_alp=alp_col_str)
            # if any of the dict entries is not empty:
            if any([len(dict_ind_rel_matched[key])>0 for key in dict_ind_rel_matched.keys()]):
                ind_valid_high_prop.extend(range(n_start, n_end))
                # set status to list of matching components
                #append all matching components from all dictionary entries to a list
                pot_cand = []
                dict_formula_by_pot_cand = {}
                dict_iupa_name_by_pot_cand = {}
                for key in dict_ind_rel_matched.keys():
                    pot_cand.extend(data_summary[key][dict_ind_rel_matched[key]])
                for pot_cand_i in pot_cand:
                    # get entries belonging to the matching components
                    if dict_ind_rel_matched[key]:
                        dict_formula_by_pot_cand[pot_cand_i] = np.array(data_summary['cs_formula'])[dict_ind_rel_matched[key]]
                    else:
                        dict_formula_by_pot_cand[pot_cand_i] = np.array([])
                    if dict_ind_rel_matched[key]:
                        dict_iupa_name_by_pot_cand[pot_cand_i] = np.array(data_summary['cs_name'])[dict_ind_rel_matched[key]]
                    else:
                        dict_iupa_name_by_pot_cand[pot_cand_i] = np.array([])
                # remove duplicates
                pot_cand = np.unique(pot_cand)
                iupac_names = [dict_iupa_name_by_pot_cand[pot_cand] for pot_cand in pot_cand]
                # flatten list
                iupac_names = [item for sublist in iupac_names for item in sublist]
                #pot_cand = np.unique(np.append(np.append(data_summary['cs_matching_cmp'][ind_rel_matched], data_summary['cs_matching_cmp'][ind_rel_matched_no_N]),data_summary['cs_matching_cmp'][ind_rel_matched_EI_CI]))
                #[formula_by_name(name, formulas_dict_manually_added = formulas_dict_manually_added, formula_dict = formula_dict) for name in pot_cand if formula_by_name(name, formulas_dict_manually_added = formulas_dict_manually_added, formula_dict = formula_dict)[3]]
                #likely_cand = []
                #for name in pot_cand:
                    #iupac_name, _, _, _ = formula_by_name(name, formulas_dict_manually_added = formulas_dict_manually_added, target_list=target_list)[1:]
                    # if char_values are not identical to the ones already in likely_cand, add them
                #    if len(iupac_name) > 0 and iupac_name not in likely_cand:
                #            likely_cand.append(iupac_name)
                if iupac_names:
                    # unique iupac names
                    iupac_names = np.unique(iupac_names)
                    print(iupac_names)
                    # append all elements of likely_cand to a string separated by ', '
                    string_likely_cand = ', '.join(iupac_names)   
                    data_summary['eval_status'][n_ind] = string_likely_cand
            else:
            # add all values between n_start and n_end ind_greyed_out
                ind_validated.extend(range(n_start, n_end))
    color_hex_table.values[ind_validated] = col_validated
    color_hex_table.values[ind_valid_high_prop] = col_validated_high_prop

    # SINGLE CELLS

    # MATCHING SCORE COLOURS
    # set green, if matching score is high enough
    # get all indici where table_data['0_score'] is not empty and >= lim_ok_score
    ind_ok = np.where(pd.to_numeric(data_summary['cs_score'], errors='coerce').fillna(-0.1) >= lim_ok_score)
    color_hex_table['cs_score'].values[ind_ok] = col_ok
    if 'cs_score_main' in data_summary.columns:
        # same for main and minor matching score
        ind_ok = np.where(pd.to_numeric(data_summary['cs_score_main'], errors='coerce').fillna(-0.1) >= lim_ok_score)
        color_hex_table['cs_score_main'].values[ind_ok] = col_ok
        # same for main minor matching score
        ind_ok = np.where(pd.to_numeric(data_summary['cs_score_minor'], errors='coerce').fillna(-0.1) >= lim_ok_score)
        color_hex_table['cs_score_minor'].values[ind_ok] = col_ok


    ind_ok = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT'], errors='coerce').fillna(1000) <= lim_ok_rt))
    color_hex_table['diff to expc. RT'].values[ind_ok] = col_ok
    if 'cs_score_main' in data_summary.columns:
        ind_ok = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT_main'], errors='coerce').fillna(1000) <= lim_ok_rt))
        color_hex_table['diff to expc. RT_main'].values[ind_ok] = col_ok
        ind_ok = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT_minor'], errors='coerce').fillna(1000) <= lim_ok_rt))
        color_hex_table['diff to expc. RT_minor'].values[ind_ok] = col_ok

    ind_alert = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT'], errors='coerce').fillna(1.1)) >=lim_alert_rt)
    color_hex_table['diff to expc. RT'].values[ind_alert] = col_alert
    if 'cs_score_main' in data_summary.columns:
        ind_alert = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT_main'], errors='coerce').fillna(1.1)) >=lim_alert_rt)
        color_hex_table['diff to expc. RT_main'].values[ind_alert] = col_alert
        ind_alert = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT_minor'], errors='coerce').fillna(1.1)) >=lim_alert_rt)
        color_hex_table['diff to expc. RT_minor'].values[ind_alert] = col_alert

    ind_careful_part1 = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT'], errors='coerce').fillna(1)) > lim_ok_rt)
    ind_careful_part2 = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT'], errors='coerce').fillna(1000)) < lim_alert_rt)
    ind_careful = np.intersect1d(ind_careful_part1, ind_careful_part2)
    color_hex_table['diff to expc. RT'].values[ind_careful] = col_careful
    if 'cs_score_main' in data_summary.columns:
        ind_careful_part1 = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT_main'], errors='coerce').fillna(1)) > lim_ok_rt)
        ind_careful_part2 = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT_main'], errors='coerce').fillna(1000)) < lim_alert_rt)
        ind_careful = np.intersect1d(ind_careful_part1, ind_careful_part2)
        color_hex_table['diff to expc. RT_main'].values[ind_careful] = col_careful
        ind_careful_part1 = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT_minor'], errors='coerce').fillna(1)) > lim_ok_rt)
        ind_careful_part2 = np.where(abs(pd.to_numeric(data_summary['diff to expc. RT_minor'], errors='coerce').fillna(1000)) < lim_alert_rt)
        ind_careful = np.intersect1d(ind_careful_part1, ind_careful_part2)
        color_hex_table['diff to expc. RT_minor'].values[ind_careful] = col_careful



    # set entries in color_hex_table['cs_matching_cmp'], color_hex_table['formula'] and color_hex_table['cs_score'] to grey, if if matching score is too low
    ind_alert = np.where(pd.to_numeric(data_summary['cs_score'], errors='coerce').fillna(1.1) < lim_alert_score)
    color_hex_table['cs_score'].values[ind_alert] = col_alert
    if 'cs_score_main' in data_summary.columns:
        ind_alert = np.where(pd.to_numeric(data_summary['cs_score_main'], errors='coerce').fillna(1.1) < lim_alert_score)
        color_hex_table['cs_score_main'].values[ind_alert] = col_alert
        ind_alert = np.where(pd.to_numeric(data_summary['cs_score_minor'], errors='coerce').fillna(1.1) < lim_alert_score)
        color_hex_table['cs_score_minor'].values[ind_alert] = col_alert

    # if below lim_alert_score set background color to col_alert
    ind_alert_part1 = np.where(pd.to_numeric(data_summary['cs_score'], errors='coerce').fillna(1.1) < lim_ok_score)
    ind_alert_part2 = np.where(pd.to_numeric(data_summary['cs_score'], errors='coerce').fillna(0.1) >= lim_alert_score)
    ind_careful = np.intersect1d(ind_alert_part1, ind_alert_part2)
    color_hex_table['cs_score'].values[ind_careful] = col_careful
    if 'cs_score_main' in data_summary.columns:
        ind_alert_part1 = np.where(pd.to_numeric(data_summary['cs_score_main'], errors='coerce').fillna(1.1) < lim_ok_score)
        ind_alert_part2 = np.where(pd.to_numeric(data_summary['cs_score_main'], errors='coerce').fillna(0.1) >= lim_alert_score)
        ind_careful = np.intersect1d(ind_alert_part1, ind_alert_part2)
        color_hex_table['cs_score_main'].values[ind_careful] = col_careful
        ind_alert_part1 = np.where(pd.to_numeric(data_summary['cs_score_minor'], errors='coerce').fillna(1.1) < lim_ok_score)
        ind_alert_part2 = np.where(pd.to_numeric(data_summary['cs_score_minor'], errors='coerce').fillna(0.1) >= lim_alert_score)
        ind_careful = np.intersect1d(ind_alert_part1, ind_alert_part2)
        color_hex_table['cs_score_minor'].values[ind_careful] = col_careful

    # NUMBER PEAKS COLOURS
    # set entries in color_hex_table['cs_matching_cmp'], color_hex_table['formula'] and color_hex_table['cs_score'] to grey, if if matching score is too low
    # indici of data_summary['alp. cmp(peaks)'] where a comma is contained in the string
    ind_careful_EI = np.where(data_summary['alp. cmp(peaks) EI'].str.contains(',')) # coeluting compounds
    ind_careful_CI = np.where(data_summary['alp. cmp(peaks) CI'].str.contains(',')) # coeluting compounds
    color_hex_table['alp. cmp(peaks) EI'].values[ind_careful_EI] = col_careful
    color_hex_table['alp. cmp(peaks) CI'].values[ind_careful_CI] = col_careful

    # if below lim_alert_score set background color to col_alert
    ind_alert = np.where(pd.to_numeric(data_summary['#peaks'], errors='coerce').fillna(np.nan) < 5)
    color_hex_table['#peaks'].values[ind_alert] = col_alert

    # all indici from ind_alert to next larger value in row_ind_not_empty are added to ind_greyout
    # ind_greyout = []
    # for n_ind in ind_alert[0]:
    #     # find next larger value in row_ind_not_empty relative to n_ind
    #     try:
    #         n_next_row_start = row_ind_not_empty[np.where(row_ind_not_empty > n_ind)[0]][0]
    #         ind_greyout.extend(range(n_ind, n_next_row_start))
    #     except:
    #         pass
    # set color of all columns of ind_greyout to grey


    # ALPINAC RESULTS COLOURS
    # set entries in color_hex_table['cs_matching_cmp'], color_hex_table['formula'] and color_hex_table['cs_score'] to grey, if if matching score is too low
    
    # if string between last ":" and first "(" of column data_summary['alpinac_results'] is identical one of the strings in column data_summary['cs_matching_cmp'][1:n]: set background color to green
    # which rows in data_summary['cmp_id'] are not empty:


    
    # find all rows where data_summary['cmp'] is not empty
    #
    #find_block(ind=0, start_rows=row_ind_not_empty)

    indici_dict = {}
    for data_col_alp in ['alpinac_EI', 'alpinac_EI_no_N', 'alpinac_CI', 'alpinac_CI_no_N']:
        ind_ok = []
        ind_ok_alp = []
        ind_subfragment = []
        ind_subfragment_alp = []
        for n_ind in range(len(data_summary[data_col_alp])):
            # find next larger value in row_ind_not_empty relative to n_ind
            try:
                n_start, _ = find_block(ind=n_ind, start_rows=row_ind_not_empty, len_table=len(data_summary))
                ind_rel_matched, ind_subf = find_alpinac_matching_pairs(n_ind=n_ind, data=data_summary, data_col_alp=data_col_alp)
                if len(ind_rel_matched) > 0:
                    ind_ok.extend(ind_rel_matched)
                    ind_ok_alp.append(n_ind)
                if len(ind_subf) > 0:
                    ind_subfragment.extend(ind_subf)
                    ind_subfragment_alp.append(n_ind) 
            except:
                pass
        indici_dict[data_col_alp] = ind_ok_alp
        indici_dict['cs_formulas'] = ind_ok
        indici_dict[data_col_alp +'_subfragment'] = ind_subfragment_alp
        indici_dict['subfragment_cs'] = ind_subfragment
        color_hex_table[data_col_alp].values[ind_ok_alp] = col_ok
        # subfragment
        color_hex_table[data_col_alp].values[ind_subfragment_alp] = col_subfragment

    # set color of indici_dict['alpinac_results'] to green
    #color_hex_table['alpinac_results'].values[indici_dict['alpinac_results' + '_subfragment']] = col_subfragment
    #color_hex_table['alpinac_results'].values[indici_dict['alpinac_results']] = col_ok
    # set color of indici_dict['alpinac_results_no_N'] to green
    #color_hex_table['alpinac_results_no_N'].values[indici_dict['alpinac_results_no_N' + '_subfragment']] = col_subfragment   
    #color_hex_table['alpinac_results_no_N'].values[indici_dict['alpinac_results_no_N']] = col_ok
    # set color of indici_dict['alpinac_results_EI_CI'] to green
    #color_hex_table['alpinac_results_EI_CI'].values[indici_dict['alpinac_results_EI_CI' + '_subfragment']] = col_subfragment
    #color_hex_table['alpinac_results_EI_CI'].values[indici_dict['alpinac_results_EI_CI']] = col_ok
    # set color of indici_dict['cs_formulas'] to green
    color_hex_table['cs_formula'].values[indici_dict['subfragment_cs']] = col_subfragment
    color_hex_table['cs_formula'].values[indici_dict['cs_formulas']] = col_ok

    return color_hex_table

def get_colour_table(data_summary):

    lim_ok_rt = 50 # upper limit
    lim_alert_rt = 200 # lower limit

    lim_ok_score = 0.75 # upper limit
    lim_alert_score = 0.5 # lower limit

    lim_ok_score_alp = 0.3
    lim_alert_score_alp = 0.1

    lim_ok_prom = 2000
    lim_alert_prom = 1000

    # Generate colour_table with the same dimensions as table_data, filled with white in hex notation
    color_hex_table = pd.DataFrame('#00FFFFFF', index=data_summary.index, columns=data_summary.columns)

    # GREY OUT UNUSABLE DATA PEAKS
    col_no = col_greyed_out

    # Create masks for the conditions
    mask_no_data = (data_summary['alp. cmp(peaks) EI'].str.startswith('no') &
                    data_summary['alp. cmp(peaks) CI'].str.startswith('no')) | \
                   (pd.to_numeric(data_summary['#peaks'], errors='coerce') < 5)

    ind_greyed_out = np.array([]).astype(int)

    for n_ind in data_summary.index:
        if mask_no_data[n_ind]:
            n_start, n_end = find_block(ind=n_ind, start_rows=row_ind_not_empty, len_table=len(data_summary))
            ind_greyed_out = np.append(ind_greyed_out, np.arange(n_start, n_end))
            data_summary.at[n_ind, 'eval_status'] = 'no usable data'

    color_hex_table.values[ind_greyed_out] = col_no

    # GREEN BLOCK IF MATCHING SCORE IS HIGH ENOUGH
    ind_validated = np.array([]).astype(int)
    ind_valid_high_prop = np.array([]).astype(int)
    alpinac_ind_match_dict = {}
    alpinac_ind_subfragment_dict = {}
    # initialize cs_formula entry in alpinac_ind_match_dict and alpinac_ind_subfragment_dict
    alpinac_ind_match_dict['cs_formula'] = []   
    alpinac_ind_subfragment_dict['cs_formula'] = []


    mask_high_score = pd.to_numeric(data_summary['cs_score'], errors='coerce') >= lim_ok_score

    for n_ind in range(len(data_summary['cs_score'])):
        if mask_high_score[n_ind]:
            n_start, n_end = find_block(ind=n_ind, start_rows=row_ind_not_empty, len_table=len(data_summary))
            dict_ind_rel_matched = {}
            dict_ind_subfragment = {}
            #find_alpinac_matching_pairs(n_ind = 392, data = data_summary, data_col_alp = 'alpinac_CI')
            for alp_col_str in ['alpinac_EI', 'alpinac_EI_no_N', 'alpinac_CI', 'alpinac_CI_no_N']:
                #initialize alpinac_ind_match_dict and alpinac_ind_subfragment_dict
                if alp_col_str not in alpinac_ind_match_dict.keys():
                    alpinac_ind_match_dict[alp_col_str] = []
                    alpinac_ind_subfragment_dict[alp_col_str] = []
                dict_ind_rel_matched[alp_col_str], dict_ind_subfragment[alp_col_str] = find_alpinac_matching_pairs(n_ind=n_ind, data=data_summary, data_col_alp=alp_col_str)
                print(n_ind, dict_ind_rel_matched, dict_ind_subfragment)
                if len(dict_ind_rel_matched[alp_col_str])>0:
                    alpinac_ind_match_dict['cs_formula'].append(list(0 + np.array(dict_ind_rel_matched[alp_col_str])))
                    alpinac_ind_match_dict[alp_col_str].append([n_ind])
                if len(dict_ind_subfragment[alp_col_str])>0:
                    alpinac_ind_subfragment_dict['cs_formula'].append(list(0 + np.array(dict_ind_subfragment[alp_col_str])))
                    alpinac_ind_subfragment_dict[alp_col_str].append([n_ind])
            if any([len(dict_ind_rel_matched[key]) > 0 for key in dict_ind_rel_matched.keys()]):
                ind_valid_high_prop = np.append(ind_valid_high_prop, np.arange(n_start, n_end))
                # add cell indices of matched compounds to dictionary
                pot_cand = []
                dict_formula_by_pot_cand = {}
                dict_iupa_name_by_pot_cand = {}
                for key in dict_ind_rel_matched.keys():
                    pot_cand_per_alpinac_mode = data_summary['cs_formula'][dict_ind_rel_matched[key]]
                    print(pot_cand_per_alpinac_mode)
                    if len(pot_cand_per_alpinac_mode) > 0:
                        print(pot_cand_per_alpinac_mode)
                        # to name string
                        #pot_cand_per_alpinac_mode_name_cleaned = [pot_cand_per_alpinac_mode_i.split(':')[-1].split('(')[0].strip() for pot_cand_per_alpinac_mode_i in pot_cand_per_alpinac_mode]
                        pot_cand = np.append(pot_cand, pot_cand_per_alpinac_mode)
                        # get unique values
                        pot_cand = np.unique(pot_cand)
                        for pot_cand_i in pot_cand:
                            dict_formula_by_pot_cand[key] = np.array(data_summary['cs_formula'])[dict_ind_rel_matched[key]]
                            dict_iupa_name_by_pot_cand[key] = np.array(data_summary['cs_name'])[dict_ind_rel_matched[key]]
                        #dict_formula_by_pot_cand[pot_cand] = np.array(data_summary['cs_formula'])[dict_ind_rel_matched[key]]
                        #dict_iupa_name_by_pot_cand[pot_cand] = np.array(data_summary['cs_name'])[dict_ind_rel_matched[key]]

                iupac_names = [dict_iupa_name_by_pot_cand[pot_cand_i] for pot_cand_i in dict_iupa_name_by_pot_cand]
                iupac_names = [item for sublist in iupac_names for item in sublist]

                if iupac_names:
                    # unique iupac names
                    iupac_names = np.unique(iupac_names)
                    string_likely_cand = ', '.join(iupac_names)
                    data_summary.at[n_ind, 'eval_status'] = string_likely_cand
            else:
                ind_validated = np.append(ind_validated, np.arange(n_start, n_end)) # add all values between n_start and n_end to only validqted (by matching score)

    color_hex_table.values[ind_validated] = col_validated
    color_hex_table.values[ind_valid_high_prop] = col_validated_high_prop



    for alp_col_str in ['alpinac_EI', 'alpinac_EI_no_N', 'alpinac_CI', 'alpinac_CI_no_N', 'cs_formula']:
         # go through all sublist in alpinac_ind_match_dict and alpinac_ind_subfragment_dict and add all elements to a list. Remove duplicates
        alpinac_ind_match_list = [item for sublist in alpinac_ind_match_dict[alp_col_str] for item in sublist]
        alpinac_ind_match_list = np.unique(alpinac_ind_match_list)
        alpinac_ind_subfragment_list = [item for sublist in alpinac_ind_subfragment_dict[alp_col_str] for item in sublist]
        alpinac_ind_subfragment_list = np.unique(alpinac_ind_subfragment_list)
        if len(alpinac_ind_subfragment_list) > 0:
            # set colour of all rows in alpinac_ind_subfragment_list to col_subfragment
            color_hex_table[alp_col_str][alpinac_ind_subfragment_list] = col_subfragment
        if len(alpinac_ind_match_list) > 0:
            # set colour of all rows in alpinac_ind_match_list to col_ok
            color_hex_table[alp_col_str][alpinac_ind_match_list] = col_ok



    # SINGLE CELLS
    def create_masks_score(data_summary, colname, lim_ok_score, lim_alert_score):
        #replace all non numeric values with np.nan
        if lim_ok_score > lim_alert_score:
            print('large is good')
            mask_ok = pd.to_numeric(data_summary[colname], errors='coerce').fillna(-0.1) >= lim_ok_score
            mask_alert = pd.to_numeric(data_summary[colname], errors='coerce').fillna(lim_ok_score + 1) < lim_alert_score
            mask_careful_part1 = pd.to_numeric(data_summary[colname], errors='coerce').fillna(-0.1) >= lim_alert_score
            mask_careful_part2 = pd.to_numeric(data_summary[colname], errors='coerce').fillna(lim_ok_score + 1) < lim_ok_score
            mask_careful = np.intersect1d(np.where(mask_careful_part1)[0], np.where(mask_careful_part2)[0])  
        else:
            print('small is good')
            #val_sign = pd.to_numeric(data_summary[frag_score_col], errors='coerce').fillna(1.1)
            #val_abs = np.abs(val_sign)
            values = pd.to_numeric(data_summary[colname], errors='coerce')
            # values replaced empty strings or \n with np.nan
            values_empty_strings_replaced = values.astype(str).replace('', np.nan)
            values_empty_strings_replaced = values_empty_strings_replaced.replace('\n', np.nan)
            values_empty_strings_replaced = values_empty_strings_replaced.replace(' ', np.nan)
            values_abs = np.abs(pd.to_numeric(values_empty_strings_replaced, errors='coerce'))
            mask_ok = values_abs.fillna(lim_ok_score + 1) <= lim_ok_score
            mask_alert = values_abs.fillna(-0.1) > lim_alert_score
            mask_careful_part1 = values_abs.fillna(lim_ok_score + 1) <= lim_alert_score
            mask_careful_part2 = values_abs.fillna(-0.1) > lim_ok_score
            mask_careful = np.intersect1d(np.where(mask_careful_part1)[0], np.where(mask_careful_part2)[0])            
        return mask_ok, mask_alert, mask_careful

    

    # MATCHING SCORE COLOURS
    # create masks for matching score
    mask_score_ok, mask_score_alert, mask_score_careful = create_masks_score(data_summary, 'cs_score', lim_ok_score, lim_alert_score)
    mask_diff_exp_rt_ok, mask_diff_exp_rt_alert, mask_diff_exp_rt_careful = create_masks_score(data_summary, 'diff to expc. RT', lim_ok_rt, lim_alert_rt)
    mask_alpinac_EI_score_ok, mask_alpinac_EI_score_alert, mask_alpinac_EI_score_careful = create_masks_score(data_summary, 'alpinac_EI_score', lim_ok_score_alp, lim_alert_score_alp)
    mask_alpinac_EI_no_N_score_ok, mask_alpinac_EI_no_N_score_alert, mask_alpinac_EI_no_N_score_careful = create_masks_score(data_summary, 'alpinac_EI_no_N_score', lim_ok_score_alp, lim_alert_score_alp)
    mask_alpinac_CI_score_ok, mask_alpinac_CI_score_alert, mask_alpinac_CI_score_careful = create_masks_score(data_summary, 'alpinac_CI_score', lim_ok_score_alp, lim_alert_score_alp)
    mask_alpinac_CI_no_N_score_ok, mask_alpinac_CI_no_N_score_alert, mask_alpinac_CI_no_N_score_careful = create_masks_score(data_summary, 'alpinac_CI_no_N_score', lim_ok_score_alp, lim_alert_score_alp)
    mask_prom_ok, mask_prom_alert, mask_prom_careful = create_masks_score(data_summary, 'prom', lim_ok_prom, lim_alert_prom)


    color_hex_table['prom'].values[mask_prom_ok] = col_ok
    color_hex_table['prom'].values[mask_prom_alert] = col_alert
    color_hex_table['prom'].values[mask_prom_careful] = col_careful

    color_hex_table['cs_score'].values[mask_score_ok] = col_ok  
    color_hex_table['cs_score'].values[mask_score_alert] = col_alert
    color_hex_table['cs_score'].values[mask_score_careful] = col_careful

    color_hex_table['diff to expc. RT'].values[mask_diff_exp_rt_ok] = col_ok
    color_hex_table['diff to expc. RT'].values[mask_diff_exp_rt_alert] = col_alert
    color_hex_table['diff to expc. RT'].values[mask_diff_exp_rt_careful] = col_careful

    color_hex_table['alpinac_EI_score'].values[mask_alpinac_EI_score_ok] = col_ok
    color_hex_table['alpinac_EI_score'].values[mask_alpinac_EI_score_alert] = col_alert
    color_hex_table['alpinac_EI_score'].values[mask_alpinac_EI_score_careful] = col_careful

    color_hex_table['alpinac_EI_no_N_score'].values[mask_alpinac_EI_no_N_score_ok] = col_ok
    color_hex_table['alpinac_EI_no_N_score'].values[mask_alpinac_EI_no_N_score_alert] = col_alert
    color_hex_table['alpinac_EI_no_N_score'].values[mask_alpinac_EI_no_N_score_careful] = col_careful

    color_hex_table['alpinac_CI_score'].values[mask_alpinac_CI_score_ok] = col_ok
    color_hex_table['alpinac_CI_score'].values[mask_alpinac_CI_score_alert] = col_alert
    color_hex_table['alpinac_CI_score'].values[mask_alpinac_CI_score_careful] = col_careful

    color_hex_table['alpinac_CI_no_N_score'].values[mask_alpinac_CI_no_N_score_ok] = col_ok
    color_hex_table['alpinac_CI_no_N_score'].values[mask_alpinac_CI_no_N_score_alert] = col_alert
    color_hex_table['alpinac_CI_no_N_score'].values[mask_alpinac_CI_no_N_score_careful] = col_careful




    #same for main and minor matching
    if 'minor_frag_score' in data_summary.columns:
        mask_score_ok_minor, mask_score_alert_minor, mask_score_careful_minor = create_masks_score(data_summary, 'minor_frag_score', lim_ok_score, lim_alert_score)
        mask_diff_exp_rt_ok_minor, mask_diff_exp_rt_alert_minor, mask_diff_exp_rt_careful_minor = create_masks_score(data_summary, 'minor_frag_rt_diff', lim_ok_rt, lim_alert_rt)

        color_hex_table['minor_frag_score'].values[mask_score_ok_minor] = col_ok
        color_hex_table['minor_frag_score'].values[mask_score_alert_minor] = col_alert
        color_hex_table['minor_frag_score'].values[mask_score_careful_minor] = col_careful

        color_hex_table['minor_frag_rt_diff'].values[mask_diff_exp_rt_ok_minor] = col_ok
        color_hex_table['minor_frag_rt_diff'].values[mask_diff_exp_rt_alert_minor] = col_alert
        color_hex_table['minor_frag_rt_diff'].values[mask_diff_exp_rt_careful_minor] = col_careful


    if 'main_frag_score' in data_summary.columns:
        mask_score_ok_main, mask_score_alert_main, mask_score_careful_main = create_masks_score(data_summary, 'main_frag_score', lim_ok_score, lim_alert_score)
        mask_diff_exp_rt_ok_main, mask_diff_exp_rt_alert_main, mask_diff_exp_rt_careful_main = create_masks_score(data_summary, 'main_frag_rt_diff', lim_ok_rt, lim_alert_rt)

        color_hex_table['main_frag_score'].values[mask_score_ok_main] = col_ok
        color_hex_table['main_frag_score'].values[mask_score_alert_main] = col_alert
        color_hex_table['main_frag_score'].values[mask_score_careful_main] = col_careful

        color_hex_table['main_frag_rt_diff'].values[mask_diff_exp_rt_ok_main] = col_ok
        color_hex_table['main_frag_rt_diff'].values[mask_diff_exp_rt_alert_main] = col_alert
        color_hex_table['main_frag_rt_diff'].values[mask_diff_exp_rt_careful_main] = col_careful


    # NUMBER PEAKS COLOURS
    mask_careful_EI = data_summary['alp. cmp(peaks) EI'].str.contains(',')
    mask_careful_CI = data_summary['alp. cmp(peaks) CI'].str.contains(',')
    mask_low_peaks = pd.to_numeric(data_summary['#peaks'], errors='coerce').fillna(np.nan) < 5

    color_hex_table['alp. cmp(peaks) EI'].values[mask_careful_EI] = col_careful
    color_hex_table['alp. cmp(peaks) CI'].values[mask_careful_CI] = col_careful
    color_hex_table['#peaks'].values[mask_low_peaks] = col_alert

    return color_hex_table



# OPTIONS
plot_summary_table = True
plot_summary_images = True
calc_summary_table = True # must be True if plot_summary_table is True or plot_single_spectra is True, otherwise no data is generated.
plot_single_spectra = True


# DEFINITIONS

# DEFINE COLOURS
from matplotlib import colors
[r_ok, g_ok, b_ok, alpha_ok] = [120/255, 240/255, 40/255, 1]
col_bg = "#00" + colors.rgb2hex((1,1,1)).replace("#","")# + "00"
col_ok = "#00" +colors.rgb2hex((r_ok, g_ok, b_ok)).replace("#","")# + "00"
col_careful ="#00" + colors.rgb2hex((1,1,0.0)).replace("#","")# + "00"
col_alert = "#00" +colors.rgb2hex((1,128/255,0)).replace("#","")# + "00"
col_subfragment = "#88" +colors.rgb2hex((0.5,0.5,1)).replace("#","")# + "00"
col_greyed_out = "#00" +colors.rgb2hex((0.5,0.5,0.5)).replace("#","")# + "00"
col_ok_alpha88 = "#88" +colors.rgb2hex((r_ok, g_ok, b_ok)).replace("#","")
col_validated = "#00" +colors.rgb2hex((0.2*1.3, 0.4*1.3, 0.1*1.3)).replace("#","")
col_validated_high_prop = "#00" +colors.rgb2hex((0.2*1.8, 0.4*1.8, 0.1*1.8)).replace("#","")

# PATHS
dir_results = Path(r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303\230401.1747.air.3")
dir_results = Path(r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303_bad_mass_calib_to_compare\230401.1747.air.3")
file_report = Path(r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303\230401.1747.air.3\230401.1747.air.3.pdf")

dir_results = Path(r"c:\Users\kaho\Desktop\data\data_Empa\Campaign202303_interesting_file\230311.0125.tank.1")
file_report = Path(r"c:\Users\kaho\Desktop\data\data_Empa\Campaign202303_interesting_file\230311.0125.tank.1\230311.0125.tank.1.report.xlsx")
file_excel_extended = dir_results / Path( "230311.0125.tank.1_peak_identification_results_extended.xlsx")

dir_results = Path(r"c:\Users\kaho\Desktop\data\data_Empa\Campaign202303_second_twin\230309.2224.tank.1")
file_report = Path(r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303_second_twin\230309.2224.tank.1\230309.2224.tank.1.report.xlsx")
file_excel_extended = dir_results / Path( "230309.2224.tank.1_peak_identification_results_extended.xlsx")


dir_results = Path(r"c:\Users\kaho\Desktop\data\data_Empa\Campaign202303_second_twin\230310.0724.tank.5")
file_report = Path(r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303_second_twin\230310.0724.tank.5\230310.0724.tank.5.report.xlsx")
file_excel_extended = dir_results / Path( "230310.0724.tank.5_peak_identification_results_extended.xlsx")





folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\Schweizerhalle_real_duplicate_2"
name_tank = '230310.1325.tank.9'

folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\BASF_real_duplicate_2"
name_tank = '230311.0125.tank.1'

folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\Solvay_real_duplicate_2"
name_tank = '230311.0555.tank.3'






folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\Solvay_single_2"
name_tank = '230311.1025.tank.5'

folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\Schweizerhalle_single_1"
name_tank = '230311.1625.tank.9'

folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\BASF_single_1"
name_tank = '230310.0724.tank.5'



name_tank = '230309.2224.tank.1'
folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\Chemours_real_duplicate_2" # seems to have wrong mass calibration

#folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\Givaudon_single_1"
#name_tank = '230310.1925.tank.11'

#folder_str = r"G:\503_Themen\Klima\kaho\results_campaign2023\Chemours_single_1" # was not calculated completely



dir_results = Path(folder_str) / Path(name_tank)
file_report = Path(folder_str) / Path(name_tank + ".report.xlsx")
file_excel_extended = dir_results / Path( name_tank + "_peak_identification_results_extended.xlsx")



# load the formula dictionary from a file in dir_results
# read out the formula dictionary
# FORMULA_DICTIONARY

# try:
#     with open(dir_results.parent / 'formula_dict.txt', 'r') as f:
#         for line in f:
#             # if line contains one ":"
#             if line.count(':') == 1:
#                 (key, val) = line.split(':')
#                 print(key)
#                 print(val)
#                 # covert string to list
#                 val = val.strip().strip('[]').split('\', ')
#                 # convert list elements to string
#                 val = [x.strip().strip("'") for x in val]
#                 # assign the list elements to the variables
#                 formula_dict[key] = val
# except:
#     print('No formula dictionary found, creating one')





# save the target list


#initiliaze Pdf class
pdf = FPDF()
# Set up the table properties
pdf.set_font('Arial', '', 8)
pdf.set_fill_color(230, 230, 230)
plot_summary_images = False
if plot_summary_images:

    pdf.add_page(orientation='P')

    # SUMMARY IMAGES
    image_path = Path(dir_results / Path(dir_results.name + "_peak_identification_results_pie_overview.png"))

    #image_path = r'C:\Users\kaho\Desktop\data\data_Empa\Campaign202303\230401.1747.air.3\230401.1747.air.3_peak_identification_results_pie_overview.png'
    # Load the image
    image = Image.open(image_path)

    # Get the size of the image in pixels
    width, height = image.size

    # Calculate the aspect ratio
    aspect_ratio = height/width

    # Calculate the width of the image in the PDF
    pdf_width = pdf.w

    # Calculate the height of the image in the PDF
    pdf_height1 = pdf_width * aspect_ratio

    # Insert the image into the PDF
    pdf.image(str(image_path), x=0, y=0, w=pdf_width, h=pdf_height1)

    #pdf.add_page(orientation='P')

    #image_path2 = r'C:\Users\kaho\Desktop\data\data_Empa\Campaign202303\230401.1747.air.3\230401.1747.air.3_peak_identification_results_pie_unidentified.png'
    image_path2 = Path(dir_results / Path(dir_results.name + "_peak_identification_results_pie_unidentified.png"))

    # Load the image
    image2 = Image.open(image_path2)

    # Get the size of the image in pixels
    width, height = image2.size

    # Calculate the aspect ratio
    aspect_ratio = height/width

    # Calculate the width of the image in the PDF
    pdf_width = pdf.w/1.2

    # Calculate the height of the image in the PDF
    pdf_height2 = pdf_width * aspect_ratio

    # Insert the image into the PDF
    pdf.image(str(image_path2), x=15, y=pdf_height1, w=pdf_width, h=pdf_height2)

    # SUMMMARY TABLE

# if plot summary table
if calc_summary_table:

    df = pd.read_excel(file_excel_extended, sheet_name='Sheet1')

    # read out header
    header = df.columns.values.tolist()

    # loop through all columns and convert those with numeric values to string values with a precision of 4 decimal places

    for col in header:
        if df[col].dtype == 'float':
            df[col] = df[col].apply(lambda x: '{0:.4g}'.format(x))

    # remove columns 'rt_start', 'rt_end', "Unamed" and 'total_intensities' and 'Status' if they exist
    for colnames_rm in ['rt_start', 'rt_end', "Unnamed: 0", 'total_intesities', 'Status']:
        if colnames_rm in df.columns:
            df = df.drop(columns=colnames_rm)


    #df = df.drop(columns=['rt_start', 'rt_end', "Unnamed: 0", 'total_intesities', 'Status'])

    # convert column name 'number_of_peaks' to '# peaks'
    df = df.rename(columns={'number_of_peaks': '#peaks'})
    df = df.rename(columns={'total_prominences': 'prom'})
    df = df.rename(columns={'extraction_number_cmps': 'alp. cmp(peaks)'})

    updated_header = df.columns.values.tolist()

    # Format the data as a table
    table_data = []
    for row in df.values:
        row_new = [str(item).replace("nan", '').replace('_nist', ' (n)').replace('_myriam',' (e)').replace('_xtra_CH', ' (x)').replace("#compound (#peaks): ","").replace("Compound_","c").replace("rank","") for item in row]
        table_data.append(list(map(str, row_new)))

    table_data = pd.DataFrame(table_data).rename(columns=dict(zip(pd.DataFrame(table_data).columns, updated_header)))
    # remove unamed index columns
    # find all columns that start with "Unnamed"
    unamed_cols = [col for col in table_data.columns if col.startswith('Unnamed')]
    table_data = table_data.drop(columns=unamed_cols)

    # set the column width to 1/2 of the page width divided by the number of columns to display
    #col_width = pdf.w / table_data.shape[1] / 2

    # calculate the column width dependent on the mamximum length of all the strings in the column and draw header cell and data cells:
    #col_width_header_list = [None]*len(updated_header)
    #col_width_data_list = [None]*len(updated_header)
    #col_width_list = [None]*len(updated_header)

    # for index, col in enumerate(updated_header):
    #     print(col)
    #     # get column col from table_data
    #     col_data = table_data[col]
    #     # get the maximum length of the string elements in the column data
    #     col_width_data_list[index] = max([pdf.get_string_width(str(elem)) for elem in list(col_data)])

    # for index, col in enumerate(updated_header):
    #     col_width_header_list[index] = max(col_width, pdf.get_string_width(col) + 2)
    #     # determine the maximum width of the string in the column data and the column header
    #     col_width_list[index] = max(col_width_header_list[index], col_width_data_list[index]) + 2

    # join content per row of columns "0_score", '1_score' and '2_score' into one string seperated by "\n" into row of column "0_score"
    table_data['0_score'] = table_data['0_score'].str.cat(table_data['1_score'], sep=';\n').str.cat(table_data['2_score'], sep=';\n')
    table_data['0_matching_cmp'] = table_data['0_matching_cmp'].str.cat(table_data['1_matching_cmp'], sep=';\n').str.cat(table_data['2_matching_cmp'], sep=';\n')
    table_data['0_name'] = table_data['0_name'].str.cat(table_data['1_name'], sep=';\n').str.cat(table_data['2_name'], sep=';\n')
    table_data['0_formula'] = table_data['0_formula'].str.cat(table_data['1_formula'], sep=';\n').str.cat(table_data['2_formula'], sep=';\n')
    table_data.insert(4, "diff to expc. RT", "")
    # calculate diff by substracting RT - RT_expected
    part_0 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['0_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
    part_1 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['1_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
    part_2 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['2_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
    table_data['diff to expc. RT'] = part_0.str.cat(part_1, sep=';\n').str.cat(part_2, sep=';\n')
    # replace nan with empty string

    # do the same for columns which are called 'main_cs_frag_0_score', 'main_cs_frag_1_score', 'main_cs_frag_2_score'
    if 'main_cs_frag_0_score' in table_data.columns:
        table_data['main_cs_frag_0_score'] = table_data['main_cs_frag_0_score'].str.cat(table_data['main_cs_frag_1_score'], sep=';\n').str.cat(table_data['main_cs_frag_2_score'], sep=';\n')
        table_data['main_cs_frag_0_matching_cmp'] = table_data['main_cs_frag_0_matching_cmp'].str.cat(table_data['main_cs_frag_1_matching_cmp'], sep=';\n').str.cat(table_data['main_cs_frag_2_matching_cmp'], sep=';\n')
        table_data['main_cs_frag_0_matching_cmp_name'] = table_data['main_cs_frag_0_matching_cmp_iupac_name'].str.cat(table_data['main_cs_frag_1_matching_cmp_iupac_name'], sep=';\n').str.cat(table_data['main_cs_frag_2_matching_cmp_iupac_name'], sep=';\n')
        table_data['main_cs_frag_0_matching_cmp_formula'] = table_data['main_cs_frag_0_matching_cmp_formula'].str.cat(table_data['main_cs_frag_1_matching_cmp_formula'], sep=';\n').str.cat(table_data['main_cs_frag_2_matching_cmp_formula'], sep=';\n')
        #table_data['main_cs_frag_0_RT_expected'] = table_data['main_cs_frag_0_matching_cmp_RT_expected'].str.cat(table_data['main_cs_frag_1_matching_cmp_RT_expected'], sep=';\n').str.cat(table_data['main_cs_frag_2_matching_cmp_RT_expected'], sep=';\n')
        # calculate diff by substracting RT - RT_expected
        part_0 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['main_cs_frag_0_matching_cmp_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
        part_1 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['main_cs_frag_1_matching_cmp_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
        part_2 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['main_cs_frag_2_matching_cmp_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
        table_data['main_cs_frag_0_matching_cmp_RT_expected'] = part_0.str.cat(part_1, sep=';\n').str.cat(part_2, sep=';\n')

    # do the same for columns which are called 'minor_cs_frag_0_score', 'minor_cs_frag_1_score', 'minor_cs_frag_2_score'
    if 'minor_cs_frag_0_score' in table_data.columns:
        table_data['minor_cs_frag_0_score'] = table_data['minor_cs_frag_0_score'].str.cat(table_data['minor_cs_frag_1_score'], sep=';\n').str.cat(table_data['minor_cs_frag_2_score'], sep=';\n')
        table_data['minor_cs_frag_0_matching_cmp'] = table_data['minor_cs_frag_0_matching_cmp'].str.cat(table_data['minor_cs_frag_1_matching_cmp'], sep=';\n').str.cat(table_data['minor_cs_frag_2_matching_cmp'], sep=';\n')
        table_data['minor_cs_frag_0_matching_cmp_name'] = table_data['minor_cs_frag_0_matching_cmp_iupac_name'].str.cat(table_data['minor_cs_frag_1_matching_cmp_iupac_name'], sep=';\n').str.cat(table_data['minor_cs_frag_2_matching_cmp_iupac_name'], sep=';\n')
        table_data['minor_cs_frag_0_matching_cmp_formula'] = table_data['minor_cs_frag_0_matching_cmp_formula'].str.cat(table_data['minor_cs_frag_1_matching_cmp_formula'], sep=';\n').str.cat(table_data['minor_cs_frag_2_matching_cmp_formula'], sep=';\n')
        #table_data['minor_cs_frag_0_matching_cmp_RT_expected'] = table_data['minor_cs_frag_0_matching_cmp_RT_expected'].str.cat(table_data['minor_cs_frag_1_matching_cmp_RT_expected'], sep=';\n').str.cat(table_data['minor_cs_frag_2_matching_cmp_RT_expected'], sep=';\n')
        # calculate diff by substracting RT - RT_expected
        part_0 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['minor_cs_frag_0_matching_cmp_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
        part_1 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['minor_cs_frag_1_matching_cmp_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
        part_2 = (pd.to_numeric(table_data['RT'], errors='coerce') - pd.to_numeric(table_data['minor_cs_frag_2_matching_cmp_RT_expected'], errors='coerce')).astype(str).replace('nan', '')
        table_data['minor_cs_frag_0_matching_cmp_RT_expected'] = part_0.str.cat(part_1, sep=';\n').str.cat(part_2, sep=';\n')
        table_data['intens_ratio_major_minor'] = (round(pd.to_numeric(table_data['main_cs_frag_intens_sum'], errors='coerce') / pd.to_numeric(table_data['minor_cs_frag_intens_sum'], errors='coerce'),ndigits = 4)).astype(str).replace('nan', '')

    table_data['0_RT_expected'] = table_data['0_RT_expected'].str.cat(table_data['1_RT_expected'], sep=';\n').str.cat(table_data['2_RT_expected'], sep=';\n')
    table_data['RT'] = table_data['RT'].str.cat(table_data['RT_min'], sep=';\n')
    # add for each elem in table_data['alp. cmp(peaks)'] the string " (#peaks)" of table_data['extraction_peaks_per_cmp_EI'] to the end of the string

    # same as for loop:
    # initialize column with empty strings
    table_data['alp. cmp(peaks) EI'] = ["" for _ in range(len(table_data['alp. cmp(peaks)']))]
    
    for ind, (elem_peaks, elem_extr) in enumerate(zip(table_data['alp. cmp(peaks)'], table_data['extraction_peaks_per_cmp_EI'])):
        table_data['alp. cmp(peaks) EI'][ind] = add_cmp_str(elem1=elem_peaks, elem2=elem_extr)

    # same for CI
    table_data['alp. cmp(peaks) CI'] = ["" for _ in range(len(table_data['alp. cmp(peaks)']))]
    for ind, (elem_peaks, elem_extr) in enumerate(zip(table_data['alp. cmp(peaks)'], table_data['extraction_peaks_per_cmp_CI'])):
        table_data['alp. cmp(peaks) CI'][ind] = add_cmp_str(elem1=elem_peaks, elem2=elem_extr)

    # initialize column with empty strings for alpinac scores



    # delete columns '1_score', '2_score', '1_matching_cmp', '2_matching_cmp'
    table_data = table_data.drop(columns=['1_score', '2_score', '1_matching_cmp', '2_matching_cmp'])
    table_data = table_data.rename(columns={'0_score': 'cs_score'})
    # rename column "0_matching_cmp" to "cs_matching_cmp"
    table_data = table_data.rename(columns={'0_matching_cmp': 'cs_matching_cmp'})
    table_data = table_data.rename(columns={'0_name': 'cs_name'})
    table_data = table_data.rename(columns={'0_formula': 'cs_formula'})
    table_data = table_data.rename(columns={'0_RT_expected': 'cs_rt'})
    # do the same for main and minor fragments
    if 'main_cs_frag_0_score' in table_data.columns:
        table_data = table_data.rename(columns={'main_cs_frag_0_score': 'main_frag_score'})
        table_data = table_data.rename(columns={'main_cs_frag_0_matching_cmp': 'main_frag_matching_cmp'})
        table_data = table_data.rename(columns={'main_cs_frag_0_matching_cmp_name': 'main_frag_name'})
        table_data = table_data.rename(columns={'main_cs_frag_0_matching_cmp_formula': 'main_frag_formula'})
        table_data = table_data.rename(columns={'main_cs_frag_0_matching_cmp_RT_expected': 'main_frag_rt_diff'})
        
    if 'minor_cs_frag_0_score' in table_data.columns:
        table_data = table_data.rename(columns={'minor_cs_frag_0_score': 'minor_frag_score'})
        table_data = table_data.rename(columns={'minor_cs_frag_0_matching_cmp': 'minor_frag_matching_cmp'})
        table_data = table_data.rename(columns={'minor_cs_frag_0_matching_cmp_name': 'minor_frag_name'})
        table_data = table_data.rename(columns={'minor_cs_frag_0_matching_cmp_formula': 'minor_frag_formula'})
        table_data = table_data.rename(columns={'minor_cs_frag_0_matching_cmp_RT_expected': 'minor_frag_rt_diff'})








    # Create a new DataFrame to store the separated rows
    #if main minor frag calc was done
    if 'main_frag_score' in table_data.columns:
        new_table_data = pd.DataFrame( columns=['ind_by_prom', 'cmp_id', 'RT', "diff to expc. RT",
        '#peaks', 'prom', 'cs_matching_cmp', 'cs_score', 'cs_name', 'cs_formula', 'cs_rt',  'alpinac_EI','alpinac_CI', 'alpinac_EI_no_N', 'alpinac_CI_no_N', 'main_cs_frag_cmp', 'main_frag_score', 'main_frag_matching_cmp', 'main_frag_name', 'main_frag_formula', 'main_frag_rt_diff', 'intens_ratio_major_minor', 'minor_cs_frag_cmp','minor_frag_score', 'minor_frag_matching_cmp', 'minor_frag_name', 'minor_frag_formula', 'minor_frag_rt_diff'])
    #if only normal calc was done
    else:
        new_table_data = pd.DataFrame( columns=['ind_by_prom', 'cmp_id', 'RT', "diff to expc. RT",
        '#peaks', 'prom', 'cs_matching_cmp', 'cs_score', 'cs_name', 'cs_formula', 'cs_rt', 'alpinac_EI','alpinac_CI', 'alpinac_EI_no_N', 'alpinac_CI_no_N'])
        # insert a column formula after col "0_matching_cmp"
    #new_table_data.insert(5, "formula", "")
    # after col "RT" insert a column "diff to expc. RT"
    #new_table_data.insert(4, "diff to expc. RT", "")
    
    #Add a column for eval_status
    #new_table_data.insert(0, "eval_status", "")

    #delete tabel_data[1_score], tabel_data[2_score], tabel_data[1_matching_cmp], tabel_data[2_matching_cmp]
    #new_table_data = new_table_data.drop(columns=['1_score', '2_score', '1_matching_cmp', '2_matching_cmp'])
    # rename column "0_score" to "cs_score"



    # Iterate over each row in the original table
    for _, row in table_data.iterrows():
        if _ == 0: row_test = row
        # Split the semicolon-separated values into a list
        scores = row['cs_score'].split(';\n')
        matching_cmps = row['cs_matching_cmp'].split(';\n')
        matching_names = row['cs_name'].split(';\n')
        matching_formulas = row['cs_formula'].split(';\n')
        rt_diffs = row["diff to expc. RT"].split(';\n')
        if 'main_frag_score' in table_data.columns:
            main_frag_scores = row['main_frag_score'].split(';\n')
            main_frag_matching_cmps = row['main_frag_matching_cmp'].split(';\n')
            main_frag_matching_names = row['main_frag_name'].split(';\n')
            main_frag_matching_formulas = row['main_frag_formula'].split(';\n')
            main_frag_rt_diffs = row['main_frag_rt_diff'].split(';\n')
            minor_frag_scores = row['minor_frag_score'].split(';\n')
            minor_frag_matching_cmps = row['minor_frag_matching_cmp'].split(';\n')
            minor_frag_matching_names = row['minor_frag_name'].split(';\n')
            minor_frag_matching_formulas = row['minor_frag_formula'].split(';\n')
            minor_frag_rt_diffs = row['minor_frag_rt_diff'].split(';\n')

        #alpinac_results_EI = row['alpinac_EI'].split(';')
        #alpinac_results_EI_no_N = row['alpinac_EI_no_N'].split(';')
        #alpinac_results_CI = row['alpinac_CI'].split(';')
        #alpinac_results_CI_no_N = row['alpinac_CI_no_N'].split(';')
        # if c0: c1: or equivalent: remove first c0: (means that evaluation of alpinac started but was cancelled in between)
        alpinac_res = [""]*4 
        for ind, col_name in enumerate(['alpinac_EI', 'alpinac_EI_no_N', 'alpinac_CI', 'alpinac_CI_no_N']):
            splitted = row[col_name].split(';')
            # while  without_not_finished is different than before: remove first element
            without_not_finished = splitted
            same = False
            while same == False:
                backup = without_not_finished   
                without_not_finished = [x.replace(x.split(':')[0]+":", "").strip() if len(x.strip())>1 and x.strip().startswith('c') and x.split(':')[1].strip().startswith('c') else x.strip() for x in without_not_finished]
                same = np.array_equal(backup, without_not_finished)
            alpinac_res[ind] = [x.replace(x.split(':')[-1].split('(')[0].strip(), chem_formula(x.split(':')[-1].split('(')[0].strip()).formula) for x in without_not_finished if x.strip() and not x.endswith(':')]
        # replace empty list by ""
        alpinac_res = [x if x else "" for x in alpinac_res]
        
        alpinac_results_EI, alpinac_results_EI_no_N, alpinac_results_CI, alpinac_results_CI_no_N = alpinac_res

        # hack because alpinac was running for too little peaks -> no candidate
        # delete list element if length of list element is < 4 and startswith 'c'
        #alpinac_results_EI_CI = [x for x in alpinac_results_EI_CI if not (len(x.strip()) <= 4 and x.strip().startswith('c'))]
  
        # remove empty strings
        #alpinac_results_EI = 
        #alpinac_results_EI_no_N = [x.replace(x.split(':')[-1].split('(')[0].strip(), chem_formula(x.split(':')[-1].split('(')[0].strip()).formula)  for x in alpinac_results_EI_no_N if x.strip()]
        #alpinac_results_CI = [x.replace(x.split(':')[-1].split('(')[0].strip(), chem_formula(x.split(':')[-1].split('(')[0].strip()).formula)  for x in alpinac_results_CI if x.strip()]
        #alpinac_results_CI_no_N = [x.replace(x.split(':')[-1].split('(')[0].strip(), chem_formula(x.split(':')[-1].split('(')[0].strip()).formula)  for x in alpinac_results_CI_no_N if x.strip()]

        #alpinac_results_EI_CI = [x.replace(x.split(':')[-1].split('(')[0].strip(), chem_formula(x.split(':')[-1].split('(')[0].strip()).formula)  for x in alpinac_results_EI_CI if x.strip()]      
        
        # replace enters by ""
        #RT_cmps = [x if x.strip() else "" for x in row['cs_rt'].split(';\n')]
        def clear_string_list(l):
            return [x if x.strip() else "" for x in l]
        scores, matching_cmps, matching_names, matching_formulas, rt_diffs = clear_string_list(scores), clear_string_list(matching_cmps), clear_string_list(matching_names), clear_string_list(matching_formulas), clear_string_list(rt_diffs)

        if 'main_frag_score' in table_data.columns:
            main_frag_scores, main_frag_matching_cmps, main_frag_matching_names, main_frag_matching_formulas, main_frag_rt_diffs = clear_string_list(main_frag_scores), clear_string_list(main_frag_matching_cmps), clear_string_list(main_frag_matching_names), clear_string_list(main_frag_matching_formulas), clear_string_list(main_frag_rt_diffs)
            minor_frag_scores, minor_frag_matching_cmps, minor_frag_matching_names, minor_frag_matching_formulas, minor_frag_rt_diffs = clear_string_list(minor_frag_scores), clear_string_list(minor_frag_matching_cmps), clear_string_list(minor_frag_matching_names), clear_string_list(minor_frag_matching_formulas), clear_string_list(minor_frag_rt_diffs)

        #RT_cmps = [x for x inif x.strip()]
        # add else = ":"

        # Check if any column contains semicolon-separated values
        if len(scores) > 1 or len(matching_cmps) > 1 or len(alpinac_results_EI) > 1 or len(alpinac_results_EI_no_N)>1 or len(alpinac_results_CI)>1 or len(alpinac_results_CI_no_N)>1:
            # Iterate over the split lists and add new rows to the new DataFrame
            for i in range(max(len(scores), len(matching_cmps), len(alpinac_results_EI), len(alpinac_results_EI_no_N), len(alpinac_results_CI), len(alpinac_results_CI_no_N))):
                new_row = row.copy()
                # set all cell values of the new_row to an empty string:
                #if new_row['RT'] != "" and i==0:
                #    RT_exp = float(new_row['RT'].split(';')[0].strip()) # get the last rt value and keep it as long as not in a new block
                if i>0:
                    new_row[:] = ''
                new_row['diff to expc. RT'] = rt_diffs[i] if i < len(rt_diffs) else ''
                new_row['cs_score'] = scores[i] if i < len(scores) else ''
                new_row['cs_matching_cmp'] = matching_cmps[i] if i < len(matching_cmps) else ''
                new_row['cs_name'] = matching_names[i] if i < len(matching_names) else ''
                new_row['cs_formula'] = matching_formulas[i] if i < len(matching_formulas) else ''
                new_row['alpinac_EI'] = alpinac_results_EI[i].split("(")[0] if i < len(alpinac_results_EI) else ''
                new_row['alpinac_EI_score'] = str(round(float(alpinac_results_EI[i].split("(")[1].split("%)")[0])/100, ndigits = 3)) if i < len(alpinac_results_EI) else ''
                new_row['alpinac_EI_no_N'] = alpinac_results_EI_no_N[i].split("(")[0] if i < len(alpinac_results_EI_no_N) else ''
                new_row['alpinac_EI_no_N_score'] = str(round(float(alpinac_results_EI_no_N[i].split("(")[1].split("%)")[0])/100, ndigits = 3)) if i < len(alpinac_results_EI_no_N) else ''
                new_row['alpinac_CI'] = alpinac_results_CI[i].split("(")[0] if i < len(alpinac_results_CI) else ''
                new_row['alpinac_CI_score'] = str(round(float(alpinac_results_CI[i].split("(")[1].split("%)")[0])/100, ndigits = 3)) if i < len(alpinac_results_CI) else ''
                new_row['alpinac_CI_no_N'] = alpinac_results_CI_no_N[i].split("(")[0] if i < len(alpinac_results_CI_no_N) else ''
                new_row['alpinac_CI_no_N_score'] = str(round(float(alpinac_results_CI_no_N[i].split("(")[1].split("%)")[0])/100, ndigits = 3)) if i < len(alpinac_results_CI_no_N) else ''

                if 'main_frag_score' in table_data.columns:
                    new_row['main_frag_score'] = main_frag_scores[i] if i < len(main_frag_scores) else ''
                    new_row['main_frag_matching_cmp'] = main_frag_matching_cmps[i] if i < len(main_frag_matching_cmps) else ''
                    new_row['main_frag_name'] = main_frag_matching_names[i] if i < len(main_frag_matching_names) else ''
                    new_row['main_frag_formula'] = main_frag_matching_formulas[i] if i < len(main_frag_matching_formulas) else ''
                    new_row['main_frag_rt_diff'] = main_frag_rt_diffs[i] if i < len(main_frag_rt_diffs) else ''
                    new_row['minor_frag_score'] = minor_frag_scores[i] if i < len(minor_frag_scores) else ''
                    new_row['minor_frag_matching_cmp'] = minor_frag_matching_cmps[i] if i < len(minor_frag_matching_cmps) else ''
                    new_row['minor_frag_name'] = minor_frag_matching_names[i] if i < len(minor_frag_matching_names) else ''
                    new_row['minor_frag_formula'] = minor_frag_matching_formulas[i] if i < len(minor_frag_matching_formulas) else ''
                    new_row['minor_frag_rt_diff'] = minor_frag_rt_diffs[i] if i < len(minor_frag_rt_diffs) else ''

                # should be done above
                # try:
                #     RT_cs_i = float(RT_cmps[i].strip())
                # except:
                #     RT_cs_i = ""
                # if RT_cs_i != "" and RT_exp != "" and i < len(matching_cmps):
                #     new_row['diff to expc. RT'] = str(RT_exp - RT_cs_i)
                # else:
                #     new_row['diff to expc. RT'] = ""


                #new_row['cs_rt_diff'] = RT_exp - int(new_row['cs_rt'].split(';')[i]) if i < len(alpinac_results_CI_no_N) and new_row['cs_rt'].split(';')[i] != "" else ''

                # fill the column formula with the formula of the compound
                #if new_row['cs_matching_cmp']:
                    #new_row['formula'], _, _, _, RT = formula_by_name(name = new_row['cs_matching_cmp'], formulas_dict_manually_added = formulas_dict_manually_added, target_list=target_list)
                    # fill the column diff to expc. RT with the difference between the experimental and the expected RT
                #    if RT:
                #new_row['diff to expc. RT'] = new_row['RT'] - new_row['cs_rt']
                # Add the new row to the new DataFrame
                new_table_data = pd.concat([new_table_data, new_row.to_frame().transpose()], ignore_index=True)
        else:
            # Add a row with empty values for columns without semicolon-separated values
            new_table_data = pd.concat([new_table_data, row.to_frame().transpose()], ignore_index=True)

    # rename column 'cs_rt_diff' to 'diff to expc. RT'
    new_table_data = new_table_data.rename(columns={'cs_rt_diff': 'diff to expc. RT'})
    # Print the new DataFrame with separated rows
    print(new_table_data)

    # add column eval_status at first position
    new_table_data.insert(0, "eval_status", "")

        # write the formula dictionary to a file for further usage
    #with open(dir_results.parent / 'formula_dict.txt', 'w') as f:
    #    for key, val in formula_dict.items():
    #        f.write('%s:%s %s' % (key, val, '\n')) # python will convert \n to os.linesep   

    # limit table data to relevant columns
    if 'main_frag_score' in new_table_data.columns:
        relevant_table_data = new_table_data[['eval_status', 'ind_by_prom', 'cmp_id', 'RT', '#peaks', 'prom', 'diff to expc. RT', 'cs_matching_cmp', 'cs_score', 'cs_name', 'cs_formula', 'alp. cmp(peaks) EI', 'alp. cmp(peaks) CI',  'alpinac_EI', 'alpinac_EI_score', 'alpinac_EI_no_N', 'alpinac_EI_no_N_score', 'alpinac_CI', 'alpinac_CI_score', 'alpinac_CI_no_N', 'alpinac_CI_no_N_score', 'main_cs_frag_cmp', 'main_frag_score', 'main_frag_matching_cmp', 'main_frag_name', 'main_frag_formula', 'main_frag_rt_diff','intens_ratio_major_minor', 'minor_cs_frag_cmp', 'minor_cs_frag_intens_sum', 'minor_frag_score', 'minor_frag_matching_cmp', 'minor_frag_name', 'minor_frag_formula', 'minor_frag_rt_diff']]
    else:
        relevant_table_data = new_table_data[['eval_status', 'ind_by_prom', 'cmp_id', 'RT', '#peaks', 'prom', 'diff to expc. RT', 'cs_matching_cmp', 'cs_score', 'cs_name', 'cs_formula', 'alp. cmp(peaks) EI', 'alp. cmp(peaks) CI', 'alpinac_EI', 'alpinac_EI_score', 'alpinac_EI_no_N', 'alpinac_EI_no_N_score', 'alpinac_CI', 'alpinac_CI_score', 'alpinac_CI_no_N', 'alpinac_CI_no_N_score']]


    # generate colour_table with same dimensions as table_data filled with white in hex notation
    colour_table = pd.DataFrame(np.zeros((relevant_table_data.shape[0], relevant_table_data.shape[1])), columns=relevant_table_data.columns.values.tolist())
    # if table_data['0_score'], table_data['1_score']  >= 0.75 set background color to green, else to white
    

    # the first line, where the component starts.
    row_ind_not_empty = np.where(relevant_table_data['cmp_id'] != '')[0] 




    # grey out all blocks for the rows in data_summary['alp. cmp(peaks)'] starts with "no"



    #color_hex_table['alpinac_results'].values[ind_greyed_out] = col_no
    #color_hex_table['alpinac_results_no_N'].values[ind_greyed_out] = col_no
    #color_hex_table['formula'].values[ind_greyed_out] = col_no
    #color_hex_table['cs_matching_cmp'].values[ind_greyed_out] = col_no
    #color_hex_table['cmp_id'].values[ind_greyed_out] = col_no
    #color_hex_table['alp. cmp(peaks)'].values[ind_greyed_out] = col_no

    colour_hex_table = get_colour_table(data_summary=relevant_table_data)

    # print colour table to excel file with colour is hex value of each cell
    #colour_hex_table.to_excel(dir_results/ "colour_hex_table.xlsx")





    # TODO: ALPINAC FORMULA ARE NOT IUPAC CONFORM! NEED TO CONVERT TO IUPAC AND THEN COMPARE


    # use data of color_hex_table to set background color of cells in table_data to the corresponding color and save it to excel file
    # TODO
    # save table_data to excel file
    #new_table_data.to_excel(dir_results/ "campaign_analysis_data_vertical_layout.xlsx")
    if 'main_frag_score' in new_table_data.columns:
        columns_to_print = ['eval_status', 'cmp_id', 'RT', 'diff to expc. RT', '#peaks', 'prom', 'cs_matching_cmp', 'cs_name', 'cs_formula', 'cs_score','alp. cmp(peaks) EI', 'alp. cmp(peaks) CI', 'alpinac_EI', 'alpinac_EI_score', 'alpinac_EI_no_N', 'alpinac_EI_no_N_score', 'alpinac_CI', 'alpinac_CI_score', 'alpinac_CI_no_N', 'alpinac_CI_no_N_score', 'main_frag_score', 'main_frag_matching_cmp', 'main_frag_name', 'main_frag_formula', 'main_frag_rt_diff', 'minor_frag_score', 'minor_frag_matching_cmp', 'minor_frag_name', 'minor_frag_formula', 'minor_frag_rt_diff']
    else:
        columns_to_print = ['eval_status', 'cmp_id', 'RT', 'diff to expc. RT', '#peaks', 'prom', 'cs_matching_cmp', 'cs_name', 'cs_formula', 'cs_score','alp. cmp(peaks) EI', 'alp. cmp(peaks) CI', 'alpinac_EI', 'alpinac_EI_score', 'alpinac_EI_no_N', 'alpinac_EI_no_N_score', 'alpinac_CI', 'alpinac_CI_score', 'alpinac_CI_no_N', 'alpinac_CI_no_N_score']
    new_table_print = new_table_data[columns_to_print]


    new_table_print.to_excel(dir_results/ (dir_results.name + "_report.xlsx"))

    #save relevant_table_data to excel file
    relevant_table_data.to_excel(dir_results/ (dir_results.name + "_report_relevant.xlsx"))



    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    # Create a new Excel workbook
    #wb = Workbook()
    #ws = wb['Sheet1']

    wb = openpyxl.load_workbook(dir_results/ (dir_results.name + "_report_relevant.xlsx"))
    ws = wb['Sheet1']

    # Write the data to the worksheet
    #pd.DataFrame(new_table_data).to_excel(wb)
    row_ind_not_empty = np.where(relevant_table_data['cmp_id'] != '')[0]
    # Apply colors to the cells
    for row in range(relevant_table_data.shape[0]):  # Start from row 2
        for col in range(relevant_table_data.shape[1]):  # Start from column B
            # get colnames
            colname = relevant_table_data.columns.values.tolist()[col]
            color_hex = colour_hex_table[colname].values[row].replace("#", "")
            #color_hex = color_hex_table.loc[row,col]
            #color_hex = '00FFFF00'
            if color_hex != '00FFFFFF':
                fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type = "lightUp")
                cell = ws.cell(row=row+2, column=col+2)
                cell.fill = fill
        # draw bold lines above row, if not empty
        if row in row_ind_not_empty:
            for col in range(relevant_table_data.shape[1]):
                cell = ws.cell(row=row+2, column=col+2)
                cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thick'))

    # Save the workbook
    wb.save(dir_results/ (dir_results.name + "_report_color.xlsx"))


    # if recalculate_matching_score_alpinac_fragments_only == True:
    #     # load the report_color.xlsx file
    #     wb = openpyxl.load_workbook(dir_results/ (dir_results.name + "_report_color.xlsx"))
    #     ws = wb['Sheet1']
    #     data = pd.DataFrame(ws.values)




    # Create a PDF object with data from and colour from colour_hex_table

# import openpyxl
# import pandas as pd
# import numpy as np
# from reportlab.lib.pagesizes import letter
# from reportlab.lib import colors
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
# from reportlab.lib.styles import getSampleStyleSheet

# # Load the workbook and select the worksheet
# wb = openpyxl.load_workbook(dir_results/ (dir_results.name + "_report_color.xlsx"))
# ws = wb['Sheet1']

# # Read the data into a DataFrame
# data = pd.DataFrame(ws.values)
# headers = data.iloc[0]
# new_table_data = pd.DataFrame(data.values[1:], columns=headers)

# # Apply colors to the cells
# colour_hex_table = pd.DataFrame(index=new_table_data.columns, columns=new_table_data.index)
# colour_hex_table.fillna('00FFFFFF', inplace=True)

# row_ind_not_empty = np.where(new_table_data['cmp_id'] != '')[0]

# # Create a PDF document
# pdf_path = Path(dir_results / "campaign_analysis_data_vertical_layout_with_colours.pdf")
# doc = SimpleDocTemplate(str(pdf_path), pagesize=letter)

# # Format the table data
# table_data = [list(headers)]
# table_data += new_table_data.values.tolist()

# # Apply colors and styles to the table
# table_style = TableStyle([
#     ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
#     ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#     ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#     ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#     ('FONTSIZE', (0, 0), (-1, 0), 12),
#     ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#     ('BACKGROUND', (0, 1), (-1, -1), colors.white),
#     ('GRID', (0, 0), (-1, -1), 1, colors.black),
# ])

# # Apply colors and styles to individual cells
# for row in range(1, len(table_data)):
#     for col in range(len(table_data[row])):
#         colname = headers[col]
#         color_hex = colour_hex_table.loc[colname, row-1].replace("#", "")
#         if color_hex != '00FFFFFF':
#             table_style.add('BACKGROUND', (col, row), (col, row), colors.HexColor(color_hex))
#         if row-1 in row_ind_not_empty:
#             table_style.add('LINEABOVE', (col, row), (col, row), 1, colors.black)

# table = Table(table_data, repeatRows=1)
# table.setStyle(table_style)

# # Build the PDF document
# elements = []
# styles = getSampleStyleSheet()
# title = Paragraph("Campaign Analysis Data", styles['Heading1'])
# elements.append(title)
# elements.append(table)
# doc.build(elements)

# print("PDF created successfully.")

STOP()

  
    










# convert the excel file to pdf
    # Add the table data
pdf.set_font('Arial', '', 8)
pdf.set_fill_color(255, 255, 255)

# calculate the column width dependent on the mamximum length of all the strings in the column:
col_width_header_list = [None]*len(updated_header)
col_width_data_list = [None]*len(updated_header)
col_width_list = [None]*len(updated_header)

def convert_to_pdf_color(wb, save_file):
    # Load the Excel file
    #excel_file = "input.xlsx"
    #wb = load_workbook(excel_file)

    # Select the active sheet
    sheet = wb.active

    # Convert the sheet to a pandas DataFrame
    data = sheet.values
    columns = next(data)[1:]

    # delete all cells outside columns
    for row in data:
        row[:] = row[0:11]
        
    




    df = pd.DataFrame(data, columns=columns)

    # Calculate the maximum character length for each column
    max_length = df.applymap(lambda x: len(str(x))).max()

    # Adjust the column widths to fit the content
    for i, col in enumerate(df.columns):
        width = max_length[col] * 1.2
        sheet.column_dimensions[chr(65 + i)].width = width

    # Create a PDF object
    pdf = FPDF(orientation="L", unit="mm", format="A4")

    # Set the font properties
    font_family = "Arial"
    font_style = "B"
    font_size = 12

    # Add a page to the PDF
    pdf.add_page()

    # Iterate over the rows and columns to add content to the PDF
    for row in pd.dataframe_to_rows(df, index=False, header=True):
        for i, value in enumerate(row):
            # Calculate the font size to fit the content within the cell
            content_width = pdf.get_string_width(str(value))
            cell_width = sheet.column_dimensions[chr(65 + i)].width
            adjusted_font_size = font_size * cell_width / content_width

            # Set the font size and style
            pdf.set_font(font_family, font_style, adjusted_font_size)

            # Add the cell content to the PDF
            pdf.cell(cell_width, font_size, str(value), border=1, ln=0)

        # Move to the next line after adding a row
        pdf.ln()

    # Save the PDF file
    pdf_file = "output.pdf"
    pdf.output(pdf_file)
    print(f"PDF file '{pdf_file}' has been created.")

convert_to_pdf_color(wb, save_file = dir_results/ "campaign_analysis_data_vertical_layout_with_colours.pdf")







STOP

# if table_data['0_matching_cmp'] is in formulas_ms set background color to green, else to white
#TODO
# save table_data to excel file
table_data.to_excel(dir_results/ "campaign_analysis_data.xlsx")

# add pdf page



# Add the table data
pdf.set_font('Arial', '', 8)
pdf.set_fill_color(255, 255, 255)

col_width = pdf.w / table_data.shape[1] / 2
row_height = pdf.font_size + 2

# get index of columns to remove
index_list = [updated_header.index("1_matching_cmp"), updated_header.index("2_matching_cmp"), updated_header.index("1_score"), updated_header.index("2_score")]
# replace col_width_list entry of each the columns ending on "_matching_cmp" with the maximum of all the col_width_list entries ending on "_matching_cmp"
col_width_list[updated_header.index("0_matching_cmp")] = max(col_width_list[updated_header.index("0_matching_cmp")], col_width_list[updated_header.index("1_matching_cmp")], col_width_list[updated_header.index("2_matching_cmp")])
col_width_list[updated_header.index("0_score")] = max(col_width_list[updated_header.index("0_score")], col_width_list[updated_header.index("1_score")], col_width_list[updated_header.index("2_score")])
# remove columns with index in index_list from data_table
table_data = table_data.drop(table_data.columns[index_list], axis=1)

# remove columns with index in index_list from col_widrh_list
col_width_list = [i for j, i in enumerate(col_width_list) if j not in index_list]

# remove columns with index in index_list from updated_header
updated_header = [i for j, i in enumerate(updated_header) if j not in index_list]

# reset col_width_list entries belonging to "alpinac_results" or "alpinac_results_no_N" to 107.08826666666667, 83.87266666666666 respectively
# TODO replot with the width found during pdf cell plotting loop on a new page and delete the page before if guess is not good
col_width_list[updated_header.index("alpinac_results")] = 50.974977777777774
col_width_list[updated_header.index("alpinac_results_no_N")] = 50.974977777777774

if plot_summary_table:

    pdf.add_page(orientation='L')
    pdf.cell(0, 10, 'Summary of measurement (50 largest compounds)', 0, 1, 'C', 1)
    pdf.ln(10)
    # Add table headers from updated_header
    for index, col in enumerate(updated_header):
        print(index)
        pdf.cell(col_width_list[index], row_height, str(col).replace("0_", "cs_"), border=1, align = 'C')
        if col == "0_matching_cmp":
            pdf.cell(col_width_list[index]/2, row_height, 'formula', border=1, align="C")
    pdf.ln()

    width_alp_res = 0
    width_alp_res_no_N = 0

    # Tidied up:
    
    # loop over the data and plot each cell
    for row_ind in range(table_data.shape[0]):
        row = [str(table_data.iloc[row_ind][key]) for key in updated_header]
        formulas_ms = ['XX']*3

        for i in range(3):
            if i==2:
                border_val = 'BLR'
            else:
                border_val = 'LR'
            for col in range(len(row)):
                elem, fill_opt, formula = get_elem(updated_header[col], row[col], i, pdf)
                print(elem)
                pdf.cell(col_width_list[col], row_height, elem, border=border_val, align="C", fill=fill_opt)
                fill_opt = 0
                if str(updated_header[col]) == "0_matching_cmp":
                    pdf.cell(col_width_list[col]/2, row_height, formula, border=border_val, align="C")
            pdf.ln(row_height)
            print(formulas_ms)



pdf.add_page(orientation='L')


# SINGLE SPECTRUM PLOTS
if plot_single_spectra:

    ind_plot = 1
    row_ind = range(ind_plot)[0]
    row_indici = [2,5]
    row_indici = [706, 610, 827, 851, 730]
    # tab

    for row_ind in row_indici:

        # define globally
        formulas_ms = ['XX']*3

        # search for libraries
        db_dir = Path(r"G:\503_Themen\Klima\TargetList")
        #df_peaks = pd.read_excel(db_dir / "TargetList.xlsx")
        df_peaks = pd.read_csv(path_target_list_org)

        # %% Load the data bases
        dbs: dict[str, dict[str, Spectrum]] = {
        "nist": JdxReader(db_dir / "nist_db").read_all(),
        "myriam": AlpinacFragmentsReader(db_dir / "myriam_ei").read_all(),
        }

        high_score_subs = table_data.iloc[row_ind]['0_matching_cmp'].replace('(n)', 'nist').replace('(e)', 'myriam').split('\n')
        database_specs = [elem.split()[0] for elem in high_score_subs]
        database_orgs = [elem.split()[1] for elem in high_score_subs]
        db_spectra = []
        db_spectra_names = []

        # get the spectra of the 3 highest scoring compounds
        for i in range(3):
            spec = dbs[database_orgs[i]][database_specs[i]]
            db_spectra.append(spec)
            db_spectra_names.append(str(database_specs[i]) + " (" + str(database_orgs[i]) + ")")


        #ind_plot = table_data.shape[0]
        #for row_ind in row_indici:
        pdf.add_page(orientation='L')
        pdf.cell(0, 10, 'Largest compound no '+str(row_ind+1), 0, 1, 'C', 1)
        pdf.ln(10)
        plot_header_and_corresponding_line(row_ind)

        # load mfg file of cmp
        mfg_file = dir_results/Path('matchms_spectra') / Path(str(table_data.iloc[row_ind]['cmp_id'])+'.mgf')
        # spectrums is a Python list of matchms-type Spectrum objects
        spectrums = list(load_from_mgf(str(mfg_file)))
        spectrums[0].metadata['compound_name'] = 'experimental'
        spectrums[0].metadata['formula'] = 'C'

        
        # process meta data, normalize intensities, reject intensities below a limit and select a certain range
        #spectrums = [metadata_processing(s) for s in spectrums]
        spectrums = [peak_processing(s) for s in spectrums]

        db_spectra = [peak_processing(s) for s in db_spectra]

        # get the spectra corresponding to the entries in the datatable
        # plot spectrum
        #spectrums.plot_mass_spectrum(spectrums[0], mirror_intensity=False, ax=axes, peak_color="C0")

        fig, axes = plt.subplots(1, 1)
        ax = axes
        # plot_spectrum(spectrums[0], mirror_intensity=False, ax=axes, peak_color="C0")
        #y_max = ax.get_ylim()[1]

        plot_mass_spectra(ax, spectrums[0], 
                            db_spectra,
                            no_plt_legend = 1)
        plt.show()
        # add logger
        logger = logging.getLogger()
        logger.setLevel(logging.DEBUG)

        path_N_yes_no = "*_EI_only_min_4_peaks_per_compound_no_N"
        path_N_yes_no = "*_EI_only_min_4_peaks_per_compound"
        path_alp_spec = dir_results #/ Path(dir_results.stem)
        # list all directories ending with  "_EI_only_min_4_peaks_per_compound_no_N"
        path_alp_spec_dirs = list(path_alp_spec.glob(path_N_yes_no))
        path_alp_spec_RT_ranges = [str(elem).split(".frag.")[1].split("s")[0:2] for elem in path_alp_spec_dirs]
        # convert string ranges to int ranges
        path_alp_spec_RT_ranges = [[int(elem[0]), int(elem[1])] for elem in path_alp_spec_RT_ranges]
        path_alp_spec_RT_mean = [np.mean(elem) for elem in path_alp_spec_RT_ranges]
        # get the RT of the experimental spectrum
        RT = int(table_data.iloc[row_ind]['RT'])
        # get the index of the RT range that is closest to the RT of the experimental spectrum
        RT_range_ind = np.argmin([abs(RT - elem) for elem in path_alp_spec_RT_mean])
        path_alp_compounds = dir_results / Path(dir_results.name + ".frag." + str(path_alp_spec_RT_ranges[RT_range_ind][0]) + "s" + str(path_alp_spec_RT_ranges[RT_range_ind][1]) + "s" +
                                                str(path_N_yes_no).split("*")[1])
        #check if fodler exists
        #if not os.path.exists(path_alp_compounds):
        #    print("Folder does not exist")
        print(path_alp_compounds)
        # list all directories starting with  "Compound":
        path_alp_compounds_dirs = list(path_alp_compounds.glob("Compound_*"))
        # get the index of the RT range that is closest to the RT of the experimental spectrum
        # resets graphics
        plt.close('all')



        if path_alp_compounds_dirs:

            alp_spec = [AlpinacData.from_alpinac(str(elem)+"/results_file_mygu.txt") for elem in path_alp_compounds_dirs]
            alp_spec_unidentified = [AlpinacData.from_alpinac_unidentified(str(elem)+"/results_file_mygu.txt") for elem in path_alp_compounds_dirs]
            alp_input_spec = [AlpinacData.from_alpinac(str(elem.parent)+".txt") for elem in path_alp_compounds_dirs]
            # path_result = r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303\230401.1747.air.3\230401.1747.air.3.frag.1835s1843s_EI_only_min_4_peaks_per_compound\Compound_0\results_file_mygu.txt"
            # path_result = r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303\230401.1747.air.3\230401.1747.air.3.frag.1735s1743s_EI_only_min_4_peaks_per_compound\Compound_2\results_file_mygu.txt"
            # alp_spec = [AlpinacData.from_alpinac(path_result)]
            # alp_spec_unidentified = [AlpinacData.from_alpinac_unidentified(path_result)]
            # alp_input_spec = [AlpinacData.from_alpinac(r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303\230401.1747.air.3\230401.1747.air.3.frag.1834s1842s.txt")]
            # alp_spec = [AlpinacData.from_alpinac(r"C:\Users\kaho\Desktop\data\data_Empa\Campaign202303\230401.1747.air.3\230401.1747.air.3.frag.1735s1743s.txt")]
            if alp_spec:
                
                fig, ax = plt.subplots(figsize = (20,10))
                for i in range(len(alp_spec)):
                    # all unidentified will be plotted at 0, limited axes to min(alp_spec[i].spectrum.intensities > 0) and max(values)
                    
                    min_ax = np.nanmin([np.nan if x == 0 else x for x in alp_spec_unidentified[i].peaks.mz])
                    ax.set_xlim([min_ax-1, max(alp_spec[i].peaks.mz)+1])

                    alp_spec[i].plot_stem(ax)
                    alp_spec_unidentified[i].plot_stem(ax, col="orange", label = ["meas. unident.", "meas. unident."])
                    # set title of figure
                    ax.set_title("File " + path_alp_compounds_dirs[0].parent.parent.name + ": " + path_alp_compounds_dirs[i].name + " of " + str(len(alp_spec)) + " in RT range " + str(path_alp_spec_RT_ranges[RT_range_ind][0]) + "s to " + str(path_alp_spec_RT_ranges[RT_range_ind][1]) + "s (Alpinac assigned fragments)")

                    # buf1 = io.BytesIO()
                    # canvas = FigureCanvas(fig)
                    # #fig.savefig(buf3, format='png')
                    # canvas.print_figure(buf1, format='png')
                    # buf1.seek(0)
                    # pil1_image = Image.open(buf1)
                add_matplotlib_plot_to_pdf(fig, pdf)

                plt.show()
            
            # # save plot to image object
            # buf = io.BytesIO()
            # canvas = FigureCanvas(fig)
            # fig.savefig(buf, format='png')
            # buf.seek(0)
            #alp_spec[0].plot(ax)
            #alp_spec[0].peaks.mz
            #alp_spec[0].peaks.intensities
            #alp_spec[0].metadata

            # add peaks of corrsponding alpinac input file before running alpinac
            # open csv file using first line as colnames:
            alp_input_raw = pd.read_csv(str(path_alp_compounds_dirs[0].parent)+".txt", sep = "\t", header = 0)
            # get column names
            alp_input_raw.columns
            # get alp_input_raw column which is called 'mass'
            
            # get mz and intensity values
            alp_input_raw['area']
            # do stem plot in grey color

            # get corresponding spectra from matchms_spectra directory
            # get the index of RT column of the excel file which is closest to RT of experimental spectrum
            RT_ind = np.argmin([abs(RT - int(elem)) for elem in table_data['RT']])
            # get cmp_id of excel file which is nearest to RT of experimental spectrum:
            cmp_id = table_data.iloc[RT_ind]['cmp_id']
            # get the path to the mgf file ending with cmp_id from matchms_spectra directory
            path_mgf_files = list(Path(dir_results/"matchms_spectra").glob("*.mgf"))
            # load mgf file ending with cmp_id from matchms_spectra directory
            mgf_file = [elem for elem in path_mgf_files if str(elem).endswith(str(cmp_id)+".mgf")][0]
            # load mgf file as spectrum
            spectrums = list(load_from_mgf(str(mgf_file)))
            # normalize spectrum
            spectrums = [(spectrum) for spectrum in spectrums]
            spectrums = [peak_processing(s) for s in spectrums]



        


            #fig, ax = plt.subplots(1, 1)
            fig, ax = plt.subplots(figsize=(20, 10))
            # set lower xlimit to 1

            plot_mass_spectra(ax, spectrums[0], col="lightgrey", label = ["py peakfinder extr.", "py peakfinder extr."])

            plot_mass_spectra(ax, alp_input_spec[0], 
                            db_spectra,
                            no_plt_legend = 1, label_peaks = "numeric")




            # set title of figure
            db_spectra_str = ""
            for ind in range(len(db_spectra_names)):
                str_add = str(ind) + ": "  + db_spectra_names[ind] + ", "
                db_spectra_str = db_spectra_str + str_add
            # reomve last comma
            db_spectra_str = db_spectra_str[:-2]    

            
            ax.set_title("File " + path_alp_compounds_dirs[0].parent.parent.name + ": " + path_alp_compounds_dirs[i].name + " of " + str(len(alp_spec)) + " in RT range " + str(path_alp_spec_RT_ranges[RT_range_ind][0]) + "s to " + str(path_alp_spec_RT_ranges[RT_range_ind][1]) + "s (Cosine similarity to database spectra: " + db_spectra_str + ")")

            buf2 = io.BytesIO()
            canvas = FigureCanvas(fig)
            #fig.savefig(buf3, format='png')
            canvas.print_figure(buf2, format='png')
            buf2.seek(0)
            pil2_image = Image.open(buf2)
            add_matplotlib_plot_to_pdf(fig, pdf)

            plt.show()



            fig, ax = plt.subplots(figsize=(20, 10))
            plot_mass_spectra(ax, alp_input_spec[0], spectrums[0], no_plt_legend = 1, label_peaks = "numeric", label = ["Alpinac extraction", "py peakfinder extr."])
            #plt.show()

            #plt.figure(figsize=(10, 10))
            #fig, ax = plt.subplots(1, 1, figsize=(20, 10), dpi = 150)
            
            # plot_spectrum(spectrums[0], mirror_intensity=False, ax=axes, peak_color="C0")
            #y_max = ax.get_ylim()[1]
            #alp_spec[0].plot_stem(ax, label_peaks = "numeric")
            ax.set_title("File " + path_alp_compounds_dirs[0].parent.parent.name + ": " + path_alp_compounds_dirs[i].name + " of " + str(len(alp_spec)) + " in RT range " + str(path_alp_spec_RT_ranges[RT_range_ind][0]) + "s to " + str(path_alp_spec_RT_ranges[RT_range_ind][1]) + "s (Alpinac (upper, calibrated) vs peakfinder (lower)")

            add_matplotlib_plot_to_pdf(fig, pdf)

            plt.show()
        
        




        buf3 = io.BytesIO()
        canvas = FigureCanvas(fig)
        #fig.savefig(buf3, format='png')
        canvas.print_figure(buf3, format='png')
        buf3.seek(0)
        pil3_image = Image.open(buf3)




        # pdf.add_page('P')


        # # add png to pdf

        # with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
        #     pil3_image.save(tmp_file.name)
        #     tmp_file.seek(0)
        #     image_path = tmp_file.name
            
        #     # Add the image to fpdf
        #     if pdf.get_y() + pil3_image.height > pdf.h:
        #         pdf.add_page('P')

        #     #shrink image if too large
        #     if pil3_image.width > 20000:
        #         pil3_image = pil3_image.resize((200, 100), Image.ANTIALIAS)
                

        #     # plot image
        #     pdf.image(image_path, x=0, y=0, w=200, h=100)
        #     # close image
        #     tmp_file.close()

        #pdf.image(pil_image, x = 0, y = 50, w = 200, h = 100)
        #pdf.image(pil2_image, x = 0, y = 150, w = 200, h = 100)
        #pdf.add_page()
        #pdf.image(pil3_image, x = 0, y = 250, w = 200, h = 100)
        



        # alp_spec_unidentified = [AlpinacData.from_alpinac_unidentified(str(path_alp_compounds)+"/results_file_mygu.txt") for elem in path_alp_compounds_dirs]
        # if alp_spec_unidentified:
        #     fig, axes = plt.subplots(1, 1)
        #     ax = axes
        #     for i in range(len(alp_spec_unidentified)):
        #         alp_spec_unidentified[i].plot(ax)
        #     plt.show()


pdf.output(file_report, 'F')
#This will generate a report PDF that includes the PNG results. You can modify the code to add more PNG images, text, or other elements to your report.

#print acutal file path
print(file_report)